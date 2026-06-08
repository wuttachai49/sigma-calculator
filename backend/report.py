"""Report generation — PDF (fpdf2) and Excel (openpyxl)."""

from __future__ import annotations
import io
import tempfile
import os
from datetime import datetime
from typing import Any

from chart import generate_opspecs_chart

# ── Shared helpers ─────────────────────────────────────────────────────────────

SIGMA_COLORS = {
    "World Class":   {"hex": "#10b981", "rgb": (16, 185, 129)},
    "Excellent":     {"hex": "#3b82f6", "rgb": (59, 130, 246)},
    "Good":          {"hex": "#f59e0b", "rgb": (245, 158, 11)},
    "Marginal":      {"hex": "#f97316", "rgb": (249, 115, 22)},
    "Unacceptable":  {"hex": "#ef4444", "rgb": (239, 68, 68)},
}

QGI_COLORS = {
    "Precision-limited": {"hex": "#6366f1", "rgb": (99, 102, 241)},
    "Mixed":             {"hex": "#f59e0b", "rgb": (245, 158, 11)},
    "Accuracy-limited":  {"hex": "#ef4444", "rgb": (239, 68, 68)},
}

COLUMNS = [
    ("Analyte",     40),
    ("Department",  28),
    ("TEa (%)",     14),
    ("Bias (%)",    14),
    ("CV (%)",      12),
    ("σ",           12),
    ("Grade",       24),
    ("QGI",         12),
    ("QGI Label",   30),
    ("QC Rules",    48),
]


# ── PDF ───────────────────────────────────────────────────────────────────────

def generate_pdf(
    results: list[dict[str, Any]],
    meta: dict[str, str] | None = None,
) -> bytes:
    from fpdf import FPDF, XPos, YPos

    def _s(text: str) -> str:
        """Encode to Latin-1, replacing unsupported chars with ASCII equivalents."""
        return (text
            .replace('–', '-').replace('—', '-')
            .replace('‘', "'").replace('’', "'")
            .replace('“', '"').replace('”', '"')
            .encode('latin-1', errors='replace').decode('latin-1'))

    meta = meta or {}
    lab      = _s(meta.get("lab", "Clinical Laboratory"))
    analyzer = _s(meta.get("analyzer", ""))
    dept     = _s(meta.get("department", ""))
    period   = _s(meta.get("period", ""))
    now      = datetime.now().strftime("%Y-%m-%d %H:%M")

    pdf = FPDF(orientation="L", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()

    # ── Title block ──────────────────────────────────────────────────────────
    pdf.set_font("Helvetica", "B", 16)
    pdf.set_fill_color(30, 30, 50)
    pdf.set_text_color(255, 255, 255)
    pdf.cell(0, 12, "Sigma Metric Report", new_x=XPos.LMARGIN, new_y=YPos.NEXT, fill=True, align="C")
    pdf.ln(1)

    pdf.set_font("Helvetica", "", 9)
    pdf.set_text_color(80, 80, 80)
    info_parts = [f"Lab: {lab}"]
    if analyzer: info_parts.append(f"Analyzer: {analyzer}")
    if dept:     info_parts.append(f"Dept: {dept}")
    if period:   info_parts.append(f"Period: {period}")
    info_parts.append(f"Generated: {now}")
    pdf.cell(0, 6, "   |   ".join(info_parts), new_x=XPos.LMARGIN, new_y=YPos.NEXT, align="C")
    pdf.ln(3)

    # ── Formula note ─────────────────────────────────────────────────────────
    pdf.set_font("Helvetica", "I", 8)
    pdf.set_text_color(120, 120, 120)
    pdf.cell(0, 5, _s("Formula: Sigma = (TEa - |Bias%|) / CV%   |   QGI = (Bias% / 1.65) / (CV% / TEa% x 0.5)"),
             new_x=XPos.LMARGIN, new_y=YPos.NEXT, align="C")
    pdf.ln(3)

    # ── Table header ─────────────────────────────────────────────────────────
    pdf.set_font("Helvetica", "B", 8)
    pdf.set_fill_color(40, 44, 68)
    pdf.set_text_color(200, 200, 220)
    for label, w in COLUMNS:
        pdf.cell(w, 7, _s(label.replace("σ", "Sigma")), border=0, align="C", fill=True)
    pdf.ln()

    # ── Table rows ────────────────────────────────────────────────────────────
    pdf.set_font("Helvetica", "", 8)
    for i, r in enumerate(results):
        bg = (245, 246, 250) if i % 2 == 0 else (255, 255, 255)
        pdf.set_fill_color(*bg)
        pdf.set_text_color(30, 30, 30)

        grade      = r.get("grade", "")
        qgi_label  = r.get("qgi_label", "")
        sc = SIGMA_COLORS.get(grade, {}).get("rgb", (100, 100, 100))
        qc = QGI_COLORS.get(qgi_label, {}).get("rgb", (100, 100, 100))

        row_h = 6
        cells = [
            (r.get("analyte", ""),                      COLUMNS[0][1], (30,30,30),   None),
            (r.get("department", ""),                   COLUMNS[1][1], (80,80,80),   None),
            (f"{r.get('tea', '')}",                     COLUMNS[2][1], (30,30,30),   None),
            (f"{r.get('bias_pct', '')}",                COLUMNS[3][1], (30,30,30),   None),
            (f"{r.get('cv_pct', '')}",                  COLUMNS[4][1], (30,30,30),   None),
            (f"{r.get('sigma', '')}",                   COLUMNS[5][1], sc,           None),
            (grade,                                     COLUMNS[6][1], sc,           None),
            (f"{r.get('qgi', '')}",                     COLUMNS[7][1], qc,           None),
            (qgi_label,                                 COLUMNS[8][1], qc,           None),
            (", ".join(x.get("plain", x.get("rule", "")) for x in r.get("qc_rules", [])), COLUMNS[9][1], (30,30,30), None),
        ]

        for text, w, color, _ in cells:
            pdf.set_text_color(*color)
            pdf.cell(w, row_h, _s(str(text)), border=0, align="C", fill=True)
        pdf.ln()

        # thin separator line
        pdf.set_draw_color(220, 220, 230)
        pdf.line(pdf.get_x(), pdf.get_y(), pdf.get_x() + sum(c[1] for c in COLUMNS), pdf.get_y())

    # ── Summary band counts ───────────────────────────────────────────────────
    pdf.ln(4)
    pdf.set_font("Helvetica", "B", 9)
    pdf.set_text_color(30, 30, 30)
    pdf.cell(0, 6, "Performance Summary", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.set_font("Helvetica", "", 8)
    from collections import Counter
    grade_counts = Counter(r.get("grade", "Unknown") for r in results)
    qgi_counts   = Counter(r.get("qgi_label", "Unknown") for r in results)
    pdf.cell(0, 5,
        _s("  Sigma bands - " + "  |  ".join(f"{g}: {n}" for g, n in grade_counts.items())),
        new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.cell(0, 5,
        _s("  QGI types   - " + "  |  ".join(f"{l}: {n}" for l, n in qgi_counts.items())),
        new_x=XPos.LMARGIN, new_y=YPos.NEXT)

    # ── QGI legend ───────────────────────────────────────────────────────────
    pdf.ln(3)
    pdf.set_font("Helvetica", "B", 8)
    pdf.cell(0, 5, "QGI Guide:", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.set_font("Helvetica", "", 7.5)
    pdf.set_text_color(80, 80, 80)
    for text in [
        "QGI < 0.8  : Precision-limited - focus on reducing CV (reagent consistency, instrument precision)",
        "QGI 0.8-1.2: Mixed - both bias and imprecision contribute - review calibration and precision sources",
        "QGI > 1.2  : Accuracy-limited  - focus on reducing Bias (calibration, matrix effects, EQA performance)",
    ]:
        pdf.cell(0, 4.5, _s(f"  * {text}"), new_x=XPos.LMARGIN, new_y=YPos.NEXT)

    # ── OPSpecs chart page ────────────────────────────────────────────────────
    chart_png = generate_opspecs_chart(results, meta, dpi=150,
                                       width_in=11.0, height_in=6.5)
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 13)
    pdf.set_text_color(30, 30, 30)
    pdf.cell(0, 10, _s("Normalized OPSpecs Chart"), new_x=XPos.LMARGIN,
             new_y=YPos.NEXT, align="C")
    pdf.set_font("Helvetica", "", 8)
    pdf.set_text_color(120, 120, 120)
    pdf.cell(0, 5,
        _s("Axes: Normalized CV% = CV / TEa x 100  |  Normalized Bias% = Bias / TEa x 100  |  Sigma lines: y = 100 - sigma x x"),
        new_x=XPos.LMARGIN, new_y=YPos.NEXT, align="C")
    pdf.ln(3)
    # Save PNG to temp file (fpdf2 requires a file path for images)
    with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as tmp:
        tmp.write(chart_png)
        tmp_path = tmp.name
    # Landscape A4: usable width ~277mm — fill most of the page
    img_w = 265
    img_h = int(img_w * 6.5 / 11.0)
    x_pos = (pdf.w - img_w) / 2
    pdf.image(tmp_path, x=x_pos, y=pdf.get_y(), w=img_w, h=img_h)
    os.unlink(tmp_path)

    return bytes(pdf.output())


# ── Excel ─────────────────────────────────────────────────────────────────────

def generate_excel(
    results: list[dict[str, Any]],
    meta: dict[str, str] | None = None,
) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, GradientFill
    )
    from openpyxl.utils import get_column_letter

    meta = meta or {}
    lab      = meta.get("lab", "Clinical Laboratory")
    analyzer = meta.get("analyzer", "")
    dept     = meta.get("department", "")
    period   = meta.get("period", "")
    now      = datetime.now().strftime("%Y-%m-%d %H:%M")

    wb = Workbook()

    # ── Sheet 1: Results ─────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Sigma Results"
    ws.sheet_view.showGridLines = False

    def _fill(hex_color: str) -> PatternFill:
        return PatternFill("solid", fgColor=hex_color.lstrip("#"))

    def _font(bold=False, color="000000", size=10):
        return Font(bold=bold, color=color, size=size, name="Calibri")

    def _align(h="center", v="center", wrap=False):
        return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

    thin = Side(style="thin", color="D0D4E0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Title row
    ws.merge_cells("A1:J1")
    ws["A1"] = "Sigma Metric Report — Clinical Laboratory Quality Management"
    ws["A1"].font      = _font(bold=True, color="FFFFFF", size=14)
    ws["A1"].fill      = _fill("1E1E32")
    ws["A1"].alignment = _align()
    ws.row_dimensions[1].height = 24

    # Meta row
    ws.merge_cells("A2:J2")
    meta_str = "   |   ".join(filter(None, [
        f"Lab: {lab}", f"Analyzer: {analyzer}" if analyzer else "",
        f"Dept: {dept}" if dept else "", f"Period: {period}" if period else "",
        f"Generated: {now}"
    ]))
    ws["A2"] = meta_str
    ws["A2"].font      = _font(color="787878", size=9)
    ws["A2"].fill      = _fill("F5F6FA")
    ws["A2"].alignment = _align()
    ws.row_dimensions[2].height = 16

    # Formula note
    ws.merge_cells("A3:J3")
    ws["A3"] = "σ = (TEa − |Bias%|) / CV%     |     QGI = (Bias% / 1.65) / (CV% / TEa% × 0.5)"
    ws["A3"].font      = _font(color="999999", size=8)
    ws["A3"].fill      = _fill("FFFFFF")
    ws["A3"].alignment = _align()
    ws.row_dimensions[3].height = 13

    # Header row
    headers = ["Analyte", "Department", "TEa (%)", "Bias (%)", "CV (%)", "σ", "Grade", "QGI", "QGI Label", "QC Rules"]
    for col, hdr in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=hdr)
        cell.font      = _font(bold=True, color="C8C8DC", size=9)
        cell.fill      = _fill("282C44")
        cell.alignment = _align()
        cell.border    = border
    ws.row_dimensions[4].height = 18

    # Data rows
    for i, r in enumerate(results):
        row_num = i + 5
        grade     = r.get("grade", "")
        qgi_label = r.get("qgi_label", "")
        sc        = SIGMA_COLORS.get(grade, {}).get("hex", "#606060").lstrip("#")
        qc        = QGI_COLORS.get(qgi_label, {}).get("hex", "#606060").lstrip("#")
        bg        = "F5F6FA" if i % 2 == 0 else "FFFFFF"

        row_vals = [
            r.get("analyte", ""),
            r.get("department", ""),
            r.get("tea", ""),
            r.get("bias_pct", ""),
            r.get("cv_pct", ""),
            r.get("sigma", ""),
            grade,
            r.get("qgi", ""),
            qgi_label,
            ", ".join(x.get("plain", x.get("rule", "")) for x in r.get("qc_rules", [])),
        ]
        color_cols = {6: sc, 7: sc, 8: qc, 9: qc}  # 1-indexed relative to row_vals

        for col, val in enumerate(row_vals, 1):
            cell = ws.cell(row=row_num, column=col, value=val)
            cell.fill      = _fill(bg)
            cell.alignment = _align(h="left" if col in (1, 2, 10) else "center", wrap=col == 10)
            cell.border    = border
            if col in color_cols:
                cell.font = _font(bold=True, color=color_cols[col], size=9)
            else:
                cell.font = _font(size=9)
        ws.row_dimensions[row_num].height = 16

    # Column widths
    col_widths = [32, 22, 10, 10, 10, 10, 18, 10, 22, 40]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Sheet 2: QGI Guide ────────────────────────────────────────────────────
    ws2 = wb.create_sheet("QGI Guide")
    ws2.sheet_view.showGridLines = False
    ws2.column_dimensions["A"].width = 18
    ws2.column_dimensions["B"].width = 22
    ws2.column_dimensions["C"].width = 60

    ws2.merge_cells("A1:C1")
    ws2["A1"] = "Quality Goal Index (QGI) Reference"
    ws2["A1"].font      = _font(bold=True, color="FFFFFF", size=12)
    ws2["A1"].fill      = _fill("1E1E32")
    ws2["A1"].alignment = _align()
    ws2.row_dimensions[1].height = 20

    ws2.merge_cells("A2:C2")
    ws2["A2"] = "QGI = (Bias% / 1.65) / (CV% / TEa% × 0.5)  — identifies whether bias or imprecision is the primary error source"
    ws2["A2"].font      = _font(color="555555", size=9)
    ws2["A2"].fill      = _fill("F5F6FA")
    ws2["A2"].alignment = _align()
    ws2.row_dimensions[2].height = 14

    for c, h in enumerate(["QGI Range", "Label", "Priority Action"], 1):
        cell = ws2.cell(row=3, column=c, value=h)
        cell.font  = _font(bold=True, color="C8C8DC", size=9)
        cell.fill  = _fill("282C44")
        cell.alignment = _align()
    ws2.row_dimensions[3].height = 16

    guide_rows = [
        ("< 0.8",   "Precision-limited", "6366F1",
         "CV dominates total error. Investigate instrument precision, reagent lot consistency, and pipetting reproducibility."),
        ("0.8–1.2", "Mixed",             "F59E0B",
         "Bias and imprecision contribute roughly equally. Review both calibration accuracy and precision sources."),
        ("> 1.2",   "Accuracy-limited",  "EF4444",
         "Bias dominates total error. Investigate calibration, reagent bias, matrix effects, and EQA performance."),
    ]
    for i, (rng, label, hex_c, action) in enumerate(guide_rows):
        row = i + 4
        c1 = ws2.cell(row=row, column=1, value=rng)
        c2 = ws2.cell(row=row, column=2, value=label)
        c3 = ws2.cell(row=row, column=3, value=action)
        for c in (c1, c2, c3):
            c.fill      = _fill("FFFFFF" if i % 2 == 0 else "F5F6FA")
            c.alignment = _align(h="left", wrap=True)
            c.border    = border
        c1.font = _font(bold=True, color=hex_c, size=9)
        c2.font = _font(bold=True, color=hex_c, size=9)
        c3.font = _font(size=9)
        ws2.row_dimensions[row].height = 28

    # ── Sheet 3: Summary ─────────────────────────────────────────────────────
    ws3 = wb.create_sheet("Summary")
    ws3.sheet_view.showGridLines = False
    ws3.column_dimensions["A"].width = 24
    ws3.column_dimensions["B"].width = 16

    ws3.merge_cells("A1:B1")
    ws3["A1"] = "Performance Summary"
    ws3["A1"].font      = _font(bold=True, color="FFFFFF", size=12)
    ws3["A1"].fill      = _fill("1E1E32")
    ws3["A1"].alignment = _align()
    ws3.row_dimensions[1].height = 20

    from collections import Counter
    grade_counts = Counter(r.get("grade", "Unknown") for r in results)
    qgi_counts   = Counter(r.get("qgi_label", "Unknown") for r in results)

    row = 2
    ws3.cell(row=row, column=1, value="Sigma Grade").font = _font(bold=True, size=9)
    ws3.cell(row=row, column=2, value="Count").font       = _font(bold=True, size=9)
    for c in (ws3.cell(row=row, column=1), ws3.cell(row=row, column=2)):
        c.fill = _fill("282C44"); c.font = _font(bold=True, color="C8C8DC", size=9)
        c.alignment = _align()

    for grade, cnt in grade_counts.most_common():
        row += 1
        hex_c = SIGMA_COLORS.get(grade, {}).get("hex", "#606060").lstrip("#")
        c1 = ws3.cell(row=row, column=1, value=grade)
        c2 = ws3.cell(row=row, column=2, value=cnt)
        c1.font = _font(bold=True, color=hex_c, size=10)
        c2.font = _font(bold=True, size=10)
        c1.alignment = _align(h="left"); c2.alignment = _align()
        ws3.row_dimensions[row].height = 16

    row += 2
    for c in (ws3.cell(row=row, column=1), ws3.cell(row=row, column=2)):
        c.value = ("QGI Label", "Count")[c.column - 1]
        c.fill  = _fill("282C44"); c.font = _font(bold=True, color="C8C8DC", size=9)
        c.alignment = _align()

    for label, cnt in qgi_counts.most_common():
        row += 1
        hex_c = QGI_COLORS.get(label, {}).get("hex", "#606060").lstrip("#")
        c1 = ws3.cell(row=row, column=1, value=label)
        c2 = ws3.cell(row=row, column=2, value=cnt)
        c1.font = _font(bold=True, color=hex_c, size=10)
        c2.font = _font(bold=True, size=10)
        c1.alignment = _align(h="left"); c2.alignment = _align()
        ws3.row_dimensions[row].height = 16

    # ── Sheet 4: OPSpecs Chart ────────────────────────────────────────────────
    try:
        from openpyxl.drawing.image import Image as XLImage

        chart_png = generate_opspecs_chart(results, meta, dpi=150,
                                           width_in=12.0, height_in=7.0)
        ws4 = wb.create_sheet("OPSpecs Chart")
        ws4.sheet_view.showGridLines = False

        ws4.merge_cells("A1:M1")
        ws4["A1"] = "Normalized OPSpecs Chart — All Analytes"
        ws4["A1"].font      = _font(bold=True, color="FFFFFF", size=13)
        ws4["A1"].fill      = _fill("1E1E32")
        ws4["A1"].alignment = _align()
        ws4.row_dimensions[1].height = 22

        ws4.merge_cells("A2:M2")
        ws4["A2"] = (
            "Axes: Normalized CV% = CV / TEa × 100   |   "
            "Normalized Bias% = Bias / TEa × 100   |   "
            "Sigma lines: y = 100 − σ × x"
        )
        ws4["A2"].font      = _font(color="888888", size=8)
        ws4["A2"].fill      = _fill("F5F6FA")
        ws4["A2"].alignment = _align()
        ws4.row_dimensions[2].height = 14

        with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as tmp:
            tmp.write(chart_png)
            tmp_path = tmp.name

        img = XLImage(tmp_path)
        img.width  = 900
        img.height = 525
        ws4.add_image(img, "A3")

        for col_letter in "ABCDEFGHIJKLM":
            ws4.column_dimensions[col_letter].width = 10
        for r in range(3, 35):
            ws4.row_dimensions[r].height = 15

        buf = io.BytesIO()
        wb.save(buf)
        os.unlink(tmp_path)  # delete AFTER save so openpyxl can read the file
        return buf.getvalue()

    except Exception:
        pass  # chart embed failed — fall through to save without chart

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
